import os
import shutil
import sys
import webbrowser
import subprocess
from contextlib import contextmanager
from urllib.parse import quote
from datetime import datetime
from pathlib import Path
import traceback

from procesos.facturas_emitidas import generar_emitidas
from services.import_a3_empresa import leer_numero_asiento_desde_a3
from services.facturae import FacturaeExporter
from procesos.facturas_word import (
    build_context_emitida,
    generar_pdf_desde_plantilla_word,
)
from utils.utilidades import (
    aplicar_descuento_total_lineas,
    get_default_output_dir,
    get_log_path,
    get_word_templates_dir,
    load_app_config,
    load_monedas,
)
from utils.validaciones import inferir_pais_desde_identificacion, normalizar_nif_cif


class FacturasEmitidasController:
    def __init__(self, gestor, codigo, ejercicio, empresa_conf, view, allow_all_years: bool = False):
        self._gestor = gestor
        self._codigo = codigo
        self._ejercicio = ejercicio
        self._empresa_conf = empresa_conf
        self._view = view
        self._allow_all_years = bool(allow_all_years)
        self._facturas_cache = []
        self._facturae_exporter = FacturaeExporter()

    @contextmanager
    def _busy_dialog(self, msg="Procesando, por favor espere..."):
        """Muestra un dialogo de espera mientras se ejecuta la operacion pesada.
        Bloquea el hilo principal pero al menos informa al usuario de que algo ocurre."""
        import tkinter as tk
        from tkinter import ttk as _ttk
        parent = None
        dlg = None
        try:
            parent = self._view.winfo_toplevel()
        except Exception:
            pass
        try:
            dlg = tk.Toplevel(parent)
            dlg.title("Gest2A3Eco")
            dlg.resizable(False, False)
            if parent:
                dlg.transient(parent)
                try:
                    parent.config(cursor="wait")
                except Exception:
                    pass
            tk.Label(dlg, text=msg, pady=14, padx=24).pack()
            pb = _ttk.Progressbar(dlg, mode="indeterminate", length=280)
            pb.pack(padx=24, pady=(0, 14))
            pb.start(10)
            dlg.update()
        except Exception:
            dlg = None
        try:
            yield
        finally:
            if dlg is not None:
                try:
                    dlg.destroy()
                except Exception:
                    pass
            if parent is not None:
                try:
                    parent.config(cursor="")
                except Exception:
                    pass

    def _security(self):
        return getattr(self._gestor, "security", None)

    def _codigo_empresa_a3(self) -> str:
        digits = "".join(ch for ch in str(self._codigo or "") if ch.isdigit())
        digits = digits.zfill(5) if digits else "00000"
        return f"E{digits[:5]}"

    def _can_write(self) -> bool:
        security = self._security()
        return True if not security else security.can_write_company(self._codigo)

    def _is_cliente(self) -> bool:
        session = getattr(self._security(), "session", None)
        return bool(session and session.role.value == "cliente")

    def _ensure_write(self, message: str | None = None) -> bool:
        if self._can_write():
            return True
        self._view.show_warning(
            "Gest2A3Eco",
            message or "Esta empresa esta en modo solo lectura para el usuario actual.",
        )
        return False

    def refresh_plantillas(self):
        pls = [p.get("nombre") for p in self._gestor.listar_emitidas(self._codigo, self._ejercicio)]
        self._view.set_plantillas(pls)

    def refresh_facturas(self):
        self._facturas_cache = self._listar_facturas_base()
        if self._allow_all_years:
            years = sorted({y for y in (self._year_from_factura(f) for f in self._facturas_cache) if y is not None})
            self._view.set_facturas_years(years)
        series = [str(f.get("serie") or "").strip() for f in self._facturas_cache if f.get("serie")]
        self._view.set_facturas_series(series)
        self.apply_facturas_filter()
        self._view.set_detalle_lineas([])

    def apply_facturas_filter(self):
        self._view.clear_facturas()
        year_filter    = self._view.get_facturas_year_filter() if self._allow_all_years else None
        serie_filter   = self._view.get_facturas_serie_filter()
        cliente_filter = self._view.get_facturas_cliente_filter()
        estado_filter  = self._view.get_facturas_estado_filter()
        for fac in self._facturas_cache:
            if year_filter is not None:
                if self._year_from_factura(fac) != year_filter:
                    continue
            if serie_filter is not None:
                if str(fac.get("serie") or "").strip() != serie_filter:
                    continue
            if cliente_filter:
                nombre = (fac.get("nombre") or "").lower()
                nif = (fac.get("nif") or "").lower()
                if cliente_filter not in nombre and cliente_filter not in nif:
                    continue
            if estado_filter is not None:
                ec = (fac.get("estado_contable") or "").strip().lower()
                # "pendiente" → sin estado contable (null/vacio en BD)
                # "contabilidad" → estado_contable='pendiente'
                # "generado" → estado_contable='generado'
                if estado_filter == "pendiente" and ec != "":
                    continue
                elif estado_filter == "contabilidad" and ec != "pendiente":
                    continue
                elif estado_filter == "generado" and ec != "generado":
                    continue
            total = self._compute_total(fac)
            self._view.insert_factura_row(fac, total)
        self._view.auto_sort_facturas()

    def refresh_albaranes(self):
        self._view.clear_albaranes()
        for alb in self._gestor.listar_albaranes_emitidas(self._codigo, self._ejercicio):
            total = self._compute_total(alb)
            self._view.insert_albaran_row(alb, total)
        self._view.auto_sort_albaranes()
        self._view.set_albaran_lineas([])

    def nueva(self):
        if not self._ensure_write():
            return
        fecha_sug = datetime.now().strftime("%d/%m/%Y")
        sugerido, serie_sug, eje_sug = self._proximo_numero_por_fecha(fecha_sug, rectificativa=False)
        cuenta_default = self._cuenta_bancaria_default()
        series_disponibles = self._listar_series_for_year(eje_sug, rectificativa=False)
        result = self._view.open_factura_dialog(
            {
                "codigo_empresa": self._codigo,
                "ejercicio": eje_sug if eje_sug is not None else self._ejercicio,
                "serie": serie_sug,
                "numero": sugerido,
                "forma_pago": "",
                "cuenta_bancaria": cuenta_default,
                "fecha_asiento": fecha_sug,
                "tipo_operacion": "01",
                "modelo_fiscal": "",
                "_series_disponibles": series_disponibles,
            },
            numero_sugerido=sugerido,
        )
        if result:
            es_borrador = bool(result.pop("_borrador", False))
            if es_borrador:
                result["numero"] = ""
                result["serie"] = serie_sug
                result["borrador"] = 1
                result["generada"] = False
                result["fecha_generacion"] = ""
                if result.get("ejercicio") is None:
                    result["ejercicio"] = self._ejercicio
                result["codigo_empresa"] = self._codigo
                self._gestor.upsert_factura_emitida(result)
            else:
                result = self._ajustar_numero_por_fecha_si_aplica(result, sugerido, serie_sug, rectificativa=False)
                result["borrador"] = 0
                result["generada"] = False
                result["fecha_generacion"] = ""
                if result.get("ejercicio") is None:
                    result["ejercicio"] = self._ejercicio
                result["codigo_empresa"] = self._codigo
                self._gestor.upsert_factura_emitida(result)
                self._incrementar_numeracion_por_factura(result, rectificativa=False)
            self.refresh_facturas()

    def nueva_rectificativa(self):
        if not self._ensure_write():
            return
        fecha_sug = datetime.now().strftime("%d/%m/%Y")
        sugerido, serie_sug, eje_sug = self._proximo_numero_por_fecha(fecha_sug, rectificativa=True)
        cuenta_default = self._cuenta_bancaria_default()
        series_disponibles = self._listar_series_for_year(eje_sug, rectificativa=True)
        result = self._view.open_factura_dialog(
            {
                "codigo_empresa": self._codigo,
                "ejercicio": eje_sug if eje_sug is not None else self._ejercicio,
                "serie": serie_sug,
                "numero": sugerido,
                "forma_pago": "",
                "cuenta_bancaria": cuenta_default,
                "fecha_asiento": fecha_sug,
                "tipo_operacion": "01",
                "modelo_fiscal": "",
                "_series_disponibles": series_disponibles,
            },
            numero_sugerido=sugerido,
        )
        if result:
            es_borrador = bool(result.pop("_borrador", False))
            if es_borrador:
                result["numero"] = ""
                result["serie"] = serie_sug
                result["borrador"] = 1
                result["generada"] = False
                result["fecha_generacion"] = ""
                if result.get("ejercicio") is None:
                    result["ejercicio"] = self._ejercicio
                result["codigo_empresa"] = self._codigo
                self._gestor.upsert_factura_emitida(result)
            else:
                result = self._ajustar_numero_por_fecha_si_aplica(result, sugerido, serie_sug, rectificativa=True)
                result["borrador"] = 0
                result["generada"] = False
                result["fecha_generacion"] = ""
                if result.get("ejercicio") is None:
                    result["ejercicio"] = self._ejercicio
                result["codigo_empresa"] = self._codigo
                self._gestor.upsert_factura_emitida(result)
                self._incrementar_numeracion_por_factura(result, rectificativa=True)
            self.refresh_facturas()

    def confirmar_borrador(self):
        """Convierte borradores seleccionados en facturas con numero asignado."""
        if not self._ensure_write():
            return
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una o mas facturas borrador.")
            return
        confirmadas = 0
        for fid in sel:
            fac = self._get_factura_by_id(fid)
            if not fac or not fac.get("borrador"):
                continue
            fecha_base = fac.get("fecha_asiento") or fac.get("fecha_expedicion") or datetime.now().strftime("%d/%m/%Y")
            serie_pref = str(fac.get("serie") or "").strip() or None
            sugerido, serie_sug, year = self._proximo_numero_por_fecha(fecha_base, rectificativa=False, nombre_serie=serie_pref)
            fac["numero"] = sugerido
            fac["serie"] = serie_sug
            fac["borrador"] = 0
            if year is not None:
                fac["ejercicio"] = year
            self._gestor.upsert_factura_emitida(fac)
            self._incrementar_numeracion_por_factura(fac, rectificativa=False)
            confirmadas += 1
        if confirmadas:
            self._view.show_info("Gest2A3Eco", f"{confirmadas} factura(s) confirmada(s) con numero asignado.")
            self.refresh_facturas()
        else:
            self._view.show_info("Gest2A3Eco", "Las facturas seleccionadas no son borradores.")

    def _check_not_generada(self, fac: dict) -> bool:
        """Devuelve True si se puede proceder. Si está generada pide contraseña."""
        if not fac.get("generada"):
            return True
        cfg = load_app_config()
        expected = str(cfg.get("desmarcar_generadas_password") or "").strip()
        if not expected:
            self._view.show_warning(
                "Gest2A3Eco",
                "Esta factura está marcada como generada.\n"
                "Configura una contraseña en Configuracion > Configurar monedas para poder editarla.",
            )
            return False
        provided = self._view.ask_desmarcar_generadas_password()
        if provided is None:
            return False
        if str(provided).strip() != expected:
            self._view.show_error("Gest2A3Eco", "Contraseña incorrecta.")
            return False
        return True

    def editar(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una factura.")
            return
        fac = self._get_factura_by_id(sel[0])
        if not fac:
            return
        if not self._check_not_generada(fac):
            return
        year = self._year_from_factura(fac) or self._ejercicio
        series_disponibles = self._listar_series_for_year(year, rectificativa=False)
        fac["_series_disponibles"] = series_disponibles
        era_borrador = bool(fac.get("borrador"))
        serie_orig = str(fac.get("serie") or "").strip()
        result = self._view.open_factura_dialog(fac)
        if result:
            es_borrador = bool(result.pop("_borrador", era_borrador))
            if era_borrador and not es_borrador:
                # Convertir borrador a factura: asignar numero
                fecha_base = result.get("fecha_asiento") or datetime.now().strftime("%d/%m/%Y")
                serie_pref = str(result.get("serie") or serie_orig or "").strip() or None
                sugerido, serie_sug, yr = self._proximo_numero_por_fecha(fecha_base, rectificativa=False, nombre_serie=serie_pref)
                result["numero"] = sugerido
                result["serie"] = serie_sug
                result["borrador"] = 0
                if yr is not None:
                    result["ejercicio"] = yr
                self._gestor.upsert_factura_emitida(result)
                self._incrementar_numeracion_por_factura(result, rectificativa=False)
            else:
                result["borrador"] = 1 if es_borrador else 0
                self._gestor.upsert_factura_emitida(result)
            self.refresh_facturas()

    def copiar(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_ids()
        if not sel:
            return
        fac = self._get_factura_by_id(sel[0])
        if not fac:
            return
        nuevo = dict(fac)
        nuevo.pop("id", None)
        nuevo["numero_asiento"] = ""
        nuevo["borrador"] = 0
        fecha_base = fac.get("fecha_asiento") or datetime.now().strftime("%d/%m/%Y")
        sugerido, serie_sug, eje_sug = self._proximo_numero_por_fecha(fecha_base, rectificativa=False)
        nuevo["numero"] = sugerido
        nuevo["serie"] = serie_sug
        nuevo["ejercicio"] = eje_sug if eje_sug is not None else self._ejercicio
        nuevo["generada"] = False
        nuevo["fecha_generacion"] = ""
        series_disponibles = self._listar_series_for_year(eje_sug, rectificativa=False)
        nuevo["_series_disponibles"] = series_disponibles
        result = self._view.open_factura_dialog(nuevo, numero_sugerido=nuevo["numero"])
        if result:
            es_borrador = bool(result.pop("_borrador", False))
            if es_borrador:
                result["numero"] = ""
                result["borrador"] = 1
                self._gestor.upsert_factura_emitida(result)
            else:
                result = self._ajustar_numero_por_fecha_si_aplica(result, sugerido, serie_sug, rectificativa=False)
                result["borrador"] = 0
                self._gestor.upsert_factura_emitida(result)
                self._incrementar_numeracion_por_factura(result, rectificativa=False)
            self.refresh_facturas()

    def rectificar(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una factura.")
            return
        fac = self._get_factura_by_id(sel[0])
        if not fac:
            return
        nuevo = dict(fac)
        nuevo.pop("id", None)
        fecha_base = fac.get("fecha_asiento") or datetime.now().strftime("%d/%m/%Y")
        sugerido, serie_sug, eje_sug = self._proximo_numero_por_fecha(fecha_base, rectificativa=True)
        nuevo["numero"] = sugerido
        nuevo["serie"] = serie_sug
        nuevo["ejercicio"] = eje_sug if eje_sug is not None else self._ejercicio
        nuevo["generada"] = False
        nuevo["fecha_generacion"] = ""
        nuevo["enviado"] = False
        nuevo["fecha_envio"] = ""
        nuevo["canal_envio"] = ""
        nuevo["borrador"] = 0
        nuevo["_allow_fecha_fuera_ejercicio"] = True
        nuevo["lineas"] = self._negate_factura_lineas(nuevo.get("lineas", []))
        nuevo["retencion_base"] = self._negate_value(nuevo.get("retencion_base"))
        nuevo["retencion_importe"] = self._negate_value(nuevo.get("retencion_importe"))
        series_rect = self._listar_series_for_year(eje_sug, rectificativa=True)
        nuevo["_series_disponibles"] = series_rect
        result = self._view.open_factura_dialog(nuevo, numero_sugerido=nuevo["numero"])
        if result:
            result.pop("_borrador", None)
            result["borrador"] = 0
            result = self._ajustar_numero_por_fecha_si_aplica(result, sugerido, serie_sug, rectificativa=True)
            self._gestor.upsert_factura_emitida(result)
            self._incrementar_numeracion_por_factura(result, rectificativa=True)
            self.refresh_facturas()

    def eliminar(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_ids()
        if not sel:
            return
        # Bloquear eliminacion de facturas generadas
        generadas_sel = [fid for fid in sel if (self._get_factura_by_id(fid) or {}).get("generada")]
        if generadas_sel:
            self._view.show_warning(
                "Gest2A3Eco",
                f"{len(generadas_sel)} factura(s) seleccionada(s) están marcadas como generadas y no pueden eliminarse.\n"
                "Desmárcalas primero para poder eliminarlas.",
            )
            return
        if not self._view.ask_yes_no("Gest2A3Eco", "Eliminar las facturas seleccionadas?"):
            return
        for fid in sel:
            fac = self._get_factura_by_id(fid)
            if not fac:
                continue
            eje = fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio
            self._gestor.eliminar_factura_emitida(self._codigo, fid, eje)
        self.refresh_facturas()

    def desmarcar_generadas(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una factura.")
            return
        cfg = load_app_config()
        expected_password = str(cfg.get("desmarcar_generadas_password") or "").strip()
        if not expected_password:
            self._view.show_warning(
                "Gest2A3Eco",
                "No hay contraseña configurada para desmarcar generadas.\nConfigúrala en Configuracion > Configurar monedas.",
            )
            return
        provided_password = self._view.ask_desmarcar_generadas_password()
        if provided_password is None:
            return
        if str(provided_password).strip() != expected_password:
            self._view.show_error("Gest2A3Eco", "Contraseña incorrecta.")
            return
        if not self._view.ask_yes_no("Gest2A3Eco", "Desmarcar como enlazadas las facturas seleccionadas?"):
            return
        if self._allow_all_years:
            grupos = {}
            for fid in sel:
                fac = self._get_factura_by_id(fid)
                if not fac:
                    continue
                eje = fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio
                grupos.setdefault(eje, []).append(fid)
            for eje, ids in grupos.items():
                self._gestor.desmarcar_facturas_emitidas_generadas(self._codigo, ids, eje)
        else:
            self._gestor.desmarcar_facturas_emitidas_generadas(self._codigo, sel, self._ejercicio)
        self.refresh_facturas()

    def factura_seleccionada(self):
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.set_detalle_lineas([])
            return
        fac = self._get_factura_by_id(sel[0])
        if not fac:
            self._view.set_detalle_lineas([])
            return
        self._view.set_detalle_lineas(fac.get("lineas", []), fac.get("moneda_simbolo", ""))

    def albaran_seleccionado(self):
        sel = self._view.get_selected_albaran_ids()
        if not sel:
            self._view.set_albaran_lineas([])
            return
        alb = self._get_albaran_by_id(sel[0])
        if not alb:
            self._view.set_albaran_lineas([])
            return
        self._view.set_albaran_lineas(alb.get("lineas", []), alb.get("moneda_simbolo", ""))

    def nuevo_albaran(self):
        if not self._ensure_write():
            return
        fecha_sug = datetime.now().strftime("%d/%m/%Y")
        sugerido, serie_sug, eje_sug = self._proximo_numero_por_fecha(fecha_sug, rectificativa=False)
        cuenta_default = self._cuenta_bancaria_default()
        result = self._view.open_albaran_dialog(
            {
                "codigo_empresa": self._codigo,
                "ejercicio": eje_sug if eje_sug is not None else self._ejercicio,
                "serie": serie_sug,
                "numero": sugerido,
                "forma_pago": "",
                "cuenta_bancaria": cuenta_default,
                "fecha_asiento": fecha_sug,
            },
            numero_sugerido=sugerido,
        )
        if result:
            result = self._ajustar_numero_por_fecha_si_aplica(result, sugerido, serie_sug, rectificativa=False)
            result["facturado"] = False
            result["factura_id"] = ""
            result["fecha_facturacion"] = ""
            if result.get("ejercicio") is None:
                result["ejercicio"] = self._ejercicio
            result["codigo_empresa"] = self._codigo
            self._gestor.upsert_albaran_emitida(result)
            self._incrementar_numeracion_por_factura(result, rectificativa=False)
            self.refresh_albaranes()

    def editar_albaran(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_albaran_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona un albaran.")
            return
        alb = self._get_albaran_by_id(sel[0])
        if not alb:
            return
        if alb.get("facturado"):
            self._view.show_warning("Gest2A3Eco", "El albaran ya esta facturado.")
            return
        result = self._view.open_albaran_dialog(alb)
        if result:
            self._gestor.upsert_albaran_emitida(result)
            self.refresh_albaranes()

    def copiar_albaran(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_albaran_ids()
        if not sel:
            return
        alb = self._get_albaran_by_id(sel[0])
        if not alb:
            return
        nuevo = dict(alb)
        nuevo.pop("id", None)
        fecha_sug = datetime.now().strftime("%d/%m/%Y")
        sugerido, serie_sug, eje_sug = self._proximo_numero_por_fecha(fecha_sug, rectificativa=False)
        nuevo["numero"] = sugerido
        nuevo["serie"] = serie_sug
        nuevo["facturado"] = False
        nuevo["factura_id"] = ""
        nuevo["fecha_facturacion"] = ""
        result = self._view.open_albaran_dialog(nuevo, numero_sugerido=sugerido)
        if result:
            result = self._ajustar_numero_por_fecha_si_aplica(result, sugerido, serie_sug, rectificativa=False)
            self._gestor.upsert_albaran_emitida(result)
            self._incrementar_numeracion_por_factura(result, rectificativa=False)
            self.refresh_albaranes()

    def eliminar_albaran(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_albaran_ids()
        if not sel:
            return
        alb = self._get_albaran_by_id(sel[0])
        if alb and alb.get("facturado"):
            self._view.show_warning("Gest2A3Eco", "El albaran ya esta facturado.")
            return
        if not self._view.ask_yes_no("Gest2A3Eco", "Eliminar los albaranes seleccionados?"):
            return
        for aid in sel:
            self._gestor.eliminar_albaran_emitida(self._codigo, aid, self._ejercicio)
        self.refresh_albaranes()

    def facturar_albaranes(self):
        if not self._ensure_write():
            return
        sel = self._view.get_selected_albaran_ids()
        if not sel:
            self._view.show_warning("Gest2A3Eco", "Selecciona uno o varios albaranes.")
            return
        try:
            if int(datetime.now().strftime("%Y")) != int(self._ejercicio):
                self._view.show_warning(
                    "Gest2A3Eco",
                    f"No se puede emitir factura fuera del ejercicio {self._ejercicio}.",
                )
                return
        except Exception:
            pass
        albaranes = []
        for aid in sel:
            alb = self._get_albaran_by_id(aid)
            if alb:
                if alb.get("facturado"):
                    continue
                albaranes.append(alb)
        if not albaranes:
            self._view.show_warning("Gest2A3Eco", "No hay albaranes pendientes de facturar.")
            return
        nifs = {normalizar_nif_cif(a.get("nif")) for a in albaranes}
        if len(nifs) > 1:
            self._view.show_warning("Gest2A3Eco", "Los albaranes seleccionados tienen distintos clientes.")
            return
        base = albaranes[0]
        moneda_codigo = base.get("moneda_codigo", "")
        moneda_simbolo = base.get("moneda_simbolo", "")
        if not moneda_codigo:
            monedas = load_monedas()
            if monedas:
                moneda_codigo = str(monedas[0].get("codigo") or "").upper()
                moneda_simbolo = str(monedas[0].get("simbolo") or "")
        descripcion_factura = self._descripcion_factura_desde_albaranes(albaranes)
        lineas = self._build_factura_lineas_desde_albaranes(albaranes)
        fecha = datetime.now().strftime("%d/%m/%Y")
        # Si es un unico albaran, la factura hereda su numero/serie (ya consumio el contador).
        # Si son varios albaranes agrupados en una factura, se usa el siguiente numero del contador.
        if len(albaranes) == 1:
            numero_fac = str(albaranes[0].get("numero") or "").strip()
            serie_fac = str(albaranes[0].get("serie") or "").strip()
            eje_fac = albaranes[0].get("ejercicio") or self._ejercicio
            incrementar_contador = False
        else:
            sugerido, serie_fac, eje_sug = self._proximo_numero_por_fecha(fecha, rectificativa=False)
            numero_fac = sugerido
            eje_fac = eje_sug if eje_sug is not None else self._ejercicio
            incrementar_contador = True
        factura = {
            "codigo_empresa": self._codigo,
            "ejercicio": eje_fac,
            "tercero_id": base.get("tercero_id"),
            "serie": serie_fac,
            "numero": numero_fac,
            "numero_largo_sii": "",
            "fecha_asiento": fecha,
            "fecha_expedicion": fecha,
            "fecha_operacion": fecha,
            "tipo_operacion": base.get("tipo_operacion") or "01",
            "modelo_fiscal": base.get("modelo_fiscal") or "",
            "nif": normalizar_nif_cif(base.get("nif", "")),
            "nombre": base.get("nombre", ""),
            "descripcion": descripcion_factura,
            "subcuenta_cliente": base.get("subcuenta_cliente", ""),
            "forma_pago": base.get("forma_pago", ""),
            "cuenta_bancaria": base.get("cuenta_bancaria", ""),
            "moneda_codigo": moneda_codigo,
            "moneda_simbolo": moneda_simbolo,
            "plantilla_word": base.get("plantilla_word") or "",
            "plantilla_emitidas": base.get("plantilla_emitidas") or self._default_plantilla_emitidas_name(eje_fac),
            "retencion_aplica": bool(base.get("retencion_aplica")),
            "retencion_pct": base.get("retencion_pct"),
            "retencion_base": base.get("retencion_base"),
            "retencion_importe": base.get("retencion_importe"),
            "lineas": lineas,
            "generada": False,
            "fecha_generacion": "",
        }
        fid = self._gestor.upsert_factura_emitida(factura)
        if incrementar_contador:
            self._incrementar_numeracion_por_factura(factura, rectificativa=False)
        fecha_gen = datetime.now().strftime("%Y-%m-%d %H:%M")
        self._gestor.marcar_albaranes_facturados(self._codigo, [a.get("id") for a in albaranes], fid, fecha_gen, self._ejercicio)
        self.refresh_facturas()
        self.refresh_albaranes()
        self._view.show_info("Gest2A3Eco", f"Factura generada desde albaranes:\n{factura.get('numero','')}")

    def imprimir_albaran(self):
        sel = self._view.get_selected_albaran_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona un albaran.")
            return
        alb = self._get_albaran_by_id(sel[0])
        if not alb:
            return

        safe_cliente = self._safe_filename(alb.get("nombre", ""))
        safe_num = self._safe_filename(alb.get("numero", ""))
        base_name = f"Albaran_{safe_num} {safe_cliente}".strip() or f"Albaran_{safe_num}"
        save_path = self._view.ask_save_pdf_path(f"{base_name}.pdf")
        if not save_path:
            return

        doc = dict(alb)
        doc.setdefault("fecha_expedicion", alb.get("fecha_asiento", ""))
        doc.setdefault("fecha_operacion", alb.get("fecha_asiento", ""))

        try:
            self._generar_pdf_word(doc, save_path, default_template="albaran_template.docx")
        except Exception as e:
            self._log_pdf_error("Error al generar PDF de albaran.", e, "", save_path)
            self._view.show_error("Gest2A3Eco", f"No se pudo generar el PDF del albaran:\n{e}")
            return

        try:
            os.startfile(save_path, "print")
            self._view.show_info("Gest2A3Eco", f"Albaran enviado a impresora:\n{save_path}")
        except Exception:
            try:
                os.startfile(save_path)
                self._view.show_info("Gest2A3Eco", f"PDF del albaran generado:\n{save_path}")
            except Exception as e:
                self._view.show_warning("Gest2A3Eco", f"PDF generado pero no se pudo abrir/imprimir:\n{e}\n\n{save_path}")

    def export_pdf(self):
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una factura.")
            return

        fac = self._get_factura_by_id(sel[0])
        if not fac:
            return

        safe_serie = self._safe_filename(str(fac.get("serie", "") or ""))
        safe_num = self._safe_filename(str(fac.get("numero", "") or ""))
        safe_codigo = self._safe_filename(self._codigo or "")
        safe_cliente = self._safe_filename(fac.get("nombre", ""))
        id_part = f"{safe_serie}{safe_num}" if (safe_serie or safe_num) else ""
        _parts = [p for p in [id_part, safe_codigo, safe_cliente] if p]
        base_name = "_".join(_parts) if _parts else f"Factura_{safe_num or fac.get('id', '')}"
        save_path = self._view.ask_save_pdf_path(f"{base_name}.pdf")
        if not save_path:
            return

        try:
            with self._busy_dialog("Generando PDF, por favor espere..."):
                self._generar_pdf_word(fac, save_path)
        except Exception as e:
            self._log_pdf_error("Error al generar PDF.", e, "", save_path)
            self._view.show_error("Gest2A3Eco", f"No se pudo generar el PDF:\n{e}")
            return

        # Guardar copia en carpeta de la aplicacion
        app_path = self._app_pdf_path(fac)
        if app_path and os.path.normpath(save_path) != os.path.normpath(app_path):
            try:
                shutil.copy2(save_path, app_path)
                upd = dict(fac)
                upd["pdf_path"] = app_path
                self._persist_factura_if_allowed(upd)
            except Exception:
                pass
        elif app_path:
            upd = dict(fac)
            upd["pdf_path"] = app_path
            self._persist_factura_if_allowed(upd)

        self._view.show_info("Gest2A3Eco", f"PDF guardado en:\n{save_path}")

    def export_pdf_multiple(self):
        import tempfile
        from pypdf import PdfWriter, PdfReader

        ids = self._view.get_marked_ids()
        if not ids:
            ids = self._view.get_selected_ids()
        if not ids:
            self._view.show_info("Gest2A3Eco", "Marca o selecciona al menos una factura.")
            return
        if len(ids) == 1:
            self.export_pdf()
            return

        save_path = self._view.ask_save_pdf_path("Facturas_seleccion.pdf")
        if not save_path:
            return

        facturas = []
        for fid in ids:
            fac = self._get_factura_by_id(fid)
            if fac:
                facturas.append(fac)
        if not facturas:
            self._view.show_error("Gest2A3Eco", "No se encontraron facturas validas.")
            return

        writer = PdfWriter()
        tmp_files = []
        errores = []
        try:
            with self._busy_dialog(f"Generando PDF ({len(facturas)} facturas), por favor espere..."):
                for fac in facturas:
                    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
                    tmp.close()
                    tmp_files.append(tmp.name)

                    try:
                        self._generar_pdf_word(fac, tmp.name)
                    except Exception as e:
                        num = fac.get("numero", fac.get("id", "?"))
                        self._log_pdf_error(f"Error al generar PDF factura {num}.", e, "", tmp.name)
                        errores.append(f"Factura {num}: {e}")
                        tmp_files.pop()
                        continue

                    try:
                        reader = PdfReader(tmp.name)
                        for page in reader.pages:
                            writer.add_page(page)
                    except Exception as e:
                        num = fac.get("numero", fac.get("id", "?"))
                        errores.append(f"Factura {num} (lectura PDF): {e}")

            if writer.pages:
                with open(save_path, "wb") as f:
                    writer.write(f)
                msg = f"PDF generado con {len(writer.pages)} pagina(s):\n{save_path}"
                if errores:
                    msg += "\n\nAdvertencias:\n" + "\n".join(errores)
                    self._view.show_warning("Gest2A3Eco", msg)
                else:
                    self._view.show_info("Gest2A3Eco", msg)
            else:
                self._view.show_error("Gest2A3Eco", "No se pudo generar ninguna pagina.\n" + "\n".join(errores))
        finally:
            for tmp_path in tmp_files:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    def generar_facturae(self):
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una factura emitida.")
            return
        fac = self._get_factura_by_id(sel[0])
        if not fac:
            return

        emisor = self._empresa_facturae()
        receptor = self._cliente_factura(fac)
        relacion = self._cliente_relacion_factura(fac)
        errores = self._facturae_exporter.validate(fac, emisor, receptor, relacion)
        if errores:
            upd = dict(fac)
            upd["facturae_status"] = "error_validacion"
            upd["facturae_error"] = "\n".join(errores)
            upd["facturae_generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self._persist_factura_if_allowed(upd)
            self.refresh_facturas()
            self._view.show_error("Gest2A3Eco", "No se puede generar Facturae/FACe:\n\n- " + "\n- ".join(errores))
            return

        nif_emisor = self._safe_filename(normalizar_nif_cif(emisor.get("nif", ""))) or "SINNIF"
        numero = self._safe_filename(self._numero_factura_contable(fac)) or "SINNUMERO"
        save_path = self._view.ask_save_xml_path(f"FACTURAE_{nif_emisor}_{numero}.xml")
        if not save_path:
            return

        result = self._facturae_exporter.export(fac, emisor, receptor, save_path, relacion)
        if not result.ok:
            upd = self._facturae_exporter.build_factura_persistence_update(fac, result)
            self._persist_factura_if_allowed(upd)
            self.refresh_facturas()
            self._view.show_error("Gest2A3Eco", "No se pudo generar el XML Facturae:\n\n- " + "\n- ".join(result.errors))
            return

        upd = self._facturae_exporter.build_factura_persistence_update(fac, result)
        self._persist_factura_if_allowed(upd)
        self.refresh_facturas()
        self._view.show_info("Gest2A3Eco", f"XML Facturae generado:\n{result.output_path}\n\n{result.warning}")
        if self._view.ask_yes_no("Gest2A3Eco", "Abrir la carpeta destino del XML Facturae?"):
            try:
                os.startfile(str(Path(result.output_path).parent))
            except Exception as exc:
                self._view.show_warning("Gest2A3Eco", f"No se pudo abrir la carpeta destino:\n{exc}")

    def abrir_pdf(self):
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una factura.")
            return
        fac = self._get_factura_by_id(sel[0])
        if not fac:
            return
        pdf_path = self._resolve_app_pdf(fac)
        if not pdf_path:
            self._view.show_warning("Gest2A3Eco", "No se pudo generar el PDF.")
            return
        try:
            os.startfile(pdf_path)
        except Exception as e:
            self._view.show_error("Gest2A3Eco", f"No se pudo abrir el PDF:\n{e}")

    def compartir_pdf(self):
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una factura.")
            return
        fac = self._get_factura_by_id(sel[0])
        if not fac:
            return
        canal = self._view.ask_share_channel()
        if not canal:
            return

        pdf_path = self._resolve_app_pdf(fac)
        if not pdf_path:
            self._view.show_warning("Gest2A3Eco", "No se pudo generar el PDF.")
            return

        related_albaranes = self._albaranes_de_factura(fac)
        attachment_paths = [pdf_path]
        include_albaranes = False
        if related_albaranes:
            cantidad = len(related_albaranes)
            texto_docs = "el albaran y la factura" if cantidad == 1 else f"los {cantidad} albaranes y la factura"
            include_albaranes = self._view.ask_yes_no(
                "Gest2A3Eco",
                f"¿Quieres enviar {texto_docs}?",
            )
            if include_albaranes:
                for alb in related_albaranes:
                    alb_pdf_path = self._resolve_albaran_pdf(alb)
                    if alb_pdf_path:
                        attachment_paths.append(alb_pdf_path)

        numero = str(fac.get("numero", "") or "")
        asunto = f"Factura {numero}".strip()
        if include_albaranes:
            cuerpo = (
                f"Adjunto factura {numero} y el albaran relacionado."
                if len(related_albaranes) == 1
                else f"Adjunto factura {numero} y los albaranes relacionados."
            ).strip()
        else:
            cuerpo = f"Adjunto factura {numero}.".strip()
        cliente = self._cliente_factura(fac)

        if canal == "email":
            from services.email_service import (
                build_html_body,
                build_outlook_bodies,
                load_email_preferences,
                load_smtp_config,
                open_outlook_email,
                save_email_preferences,
                save_smtp_config,
                send_email_smtp,
            )
            email_cliente = str(cliente.get("email") or "").strip()
            email_empresa = str(self._empresa_conf.get("email") or "").strip()
            email_prefs = load_email_preferences()
            email_mode = str(email_prefs.get("email_mode") or "outlook").strip().lower() or "outlook"
            smtp_cfg = load_smtp_config() if email_mode == "smtp" else {}

            if email_mode == "smtp" and not smtp_cfg.get("host"):
                smtp_cfg = self._view.ask_smtp_config(smtp_cfg or {})
                if not smtp_cfg or not smtp_cfg.get("host"):
                    return
                save_smtp_config(smtp_cfg)

            compose = self._view.ask_email_compose(
                email_cliente, asunto, cuerpo, pdf_path, smtp_cfg,
                email_empresa=email_empresa,
                email_mode=email_mode,
                default_cc=str(email_prefs.get("default_cc") or ""),
                default_bcc=str(email_prefs.get("default_bcc") or ""),
                email_signature=str(email_prefs.get("email_signature") or ""),
            )
            if not compose:
                return

            tot = self._totales_factura(fac)
            html_body = build_html_body(self._empresa_conf, fac, cliente, tot)
            plain_body, outlook_html_body = build_outlook_bodies(
                compose["cuerpo"],
                html_body=html_body,
                signature=compose.get("signature") or "",
            )

            save_email_preferences(
                {
                    "email_mode": email_mode,
                    "default_cc": compose.get("cc", ""),
                    "default_bcc": compose.get("bcc", ""),
                    "email_signature": compose.get("signature", ""),
                    "open_outlook_before_send": True,
                }
            )

            if email_mode == "smtp":
                new_smtp = compose.get("smtp_cfg") or {}
                if new_smtp and new_smtp != smtp_cfg:
                    save_smtp_config(new_smtp)
                    smtp_cfg = new_smtp
                try:
                    send_email_smtp(
                        smtp_cfg, compose["emails"], compose["asunto"], plain_body,
                        pdf_path, attachment_paths=attachment_paths[1:], html_body=outlook_html_body,
                    )
                    self._view.show_info("Gest2A3Eco", "Email enviado correctamente.")
                except Exception as e:
                    host = str(smtp_cfg.get("host") or "(vacio)").strip() or "(vacio)"
                    port = smtp_cfg.get("port", 587)
                    self._view.show_error(
                        "Gest2A3Eco",
                        f"Error al enviar el email:\n{e}\n\n"
                        f"Servidor configurado: {host}:{port}\n"
                        f"Verifica la configuracion SMTP pulsando 'Cambiar SMTP' en el dialogo de envio.",
                    )
                    return
            else:
                try:
                    open_outlook_email(
                        to="; ".join(compose["emails"]),
                        subject=compose["asunto"],
                        body=plain_body,
                        attachments=attachment_paths,
                        cc=compose.get("cc", ""),
                        bcc=compose.get("bcc", ""),
                        html_body=outlook_html_body,
                    )
                    self._view.show_info(
                        "Gest2A3Eco",
                        "Se ha abierto Outlook con el correo preparado. Revise el mensaje y pulse Enviar desde Outlook.",
                    )
                except FileNotFoundError as e:
                    self._view.show_error("Gest2A3Eco", str(e))
                    return
                except Exception as e:
                    self._view.show_error(
                        "Gest2A3Eco",
                        f"{e}\n\nEl correo no se ha enviado y la aplicacion puede seguir utilizandose.",
                    )
                    return

        elif canal == "whatsapp":
            telefono_cliente = str(cliente.get("telefono") or "").strip()
            telefono_empresa = str(self._empresa_conf.get("telefono") or "").strip()
            compose_wa = self._view.ask_whatsapp_compose(
                telefono_cliente, cuerpo, pdf_path,
                telefono_empresa=telefono_empresa,
            )
            if not compose_wa:
                return
            tel = compose_wa["telefono"]
            msg = compose_wa["mensaje"]
            from urllib.parse import quote
            url = f"https://wa.me/{tel}?text={quote(msg)}" if tel else "https://web.whatsapp.com/"
            try:
                webbrowser.open(url)
            except Exception:
                pass
            try:
                self._open_attachments_in_explorer(attachment_paths)
            except Exception:
                pass

        if self._view.ask_yes_no("Gest2A3Eco", "¿Marcar factura como enviada?"):
            if not self._ensure_write("Necesitas permiso de escritura para marcar la factura como enviada."):
                return
            fecha_envio = datetime.now().strftime("%Y-%m-%d %H:%M")
            eje = fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio
            self._gestor.marcar_factura_emitida_enviada(
                self._codigo, fac.get("id"), fecha_envio, canal, eje
            )
            self.refresh_facturas()

    def generar_suenlace(self):
        if not self._ensure_write():
            return
        sel = self._view.get_marked_ids()
        if not sel:
            sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_warning("Gest2A3Eco", "Marca o selecciona al menos una factura.")
            return
        facturas_sel = []
        for fid in sel:
            fac = self._get_factura_by_id(fid)
            if fac:
                facturas_sel.append(fac)

        ya_generadas = [f for f in facturas_sel if f.get("generada")]
        if ya_generadas:
            nums = ", ".join(f.get('numero','') for f in ya_generadas)
            if not self._view.ask_yes_no(
                "Gest2A3Eco",
                "Las facturas {} ya estan marcadas como generadas.\nGenerar suenlace de todas formas?".format(nums),
            ):
                return

        ndig = int(self._empresa_conf.get("digitos_plan") or 8)
        terceros = self._gestor.listar_terceros()
        terceros_by_nif = {
            normalizar_nif_cif(t.get("nif")): t
            for t in terceros
            if normalizar_nif_cif(t.get("nif"))
        }
        terceros_empresa = self._gestor.listar_terceros_por_empresa(self._codigo, self._ejercicio)
        for t in terceros_empresa:
            nif = normalizar_nif_cif(t.get("nif"))
            if nif:
                terceros_by_nif[nif] = t
        plantillas_cache = {}
        no_tpl_years = set()
        registros = []
        with self._busy_dialog("Generando suenlace.dat, por favor espere..."):
            facturas_sel = self._prepare_facturas_for_suenlace(facturas_sel)
            for fac in facturas_sel:
                rows = self._factura_to_rows(fac)
                if not rows:
                    continue
                plantilla = self._plantilla_emitidas_for_factura(fac, plantillas_cache, no_tpl_years)
                registros.extend(
                    generar_emitidas(
                        rows,
                        plantilla,
                        str(self._codigo),
                        ndig,
                        ejercicio=self._ejercicio,
                        terceros_by_nif=terceros_by_nif,
                        formato_512=True,
                    )
                )
        if no_tpl_years:
            years = ", ".join(str(x) for x in sorted(no_tpl_years))
            self._view.show_warning(
                "Gest2A3Eco",
                "No hay plantillas de emitidas para los ejercicios:\n"
                f"{years}\nSe generara usando la plantilla tipo de facturas emitidas.",
            )
        if not registros:
            self._view.show_warning("Gest2A3Eco", "No se generaron registros.")
            return

        save_path = self._view.ask_save_dat_path(f"{self._codigo_empresa_a3()}.dat")
        if not save_path:
            return
        with open(save_path, "w", encoding="latin-1", newline="") as f:
            f.writelines(registros)
        fecha_gen = datetime.now().strftime("%Y-%m-%d %H:%M")
        if self._allow_all_years:
            grupos = {}
            for fac in facturas_sel:
                eje = fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio
                grupos.setdefault(eje, []).append(fac.get("id"))
            for eje, ids in grupos.items():
                self._gestor.marcar_facturas_emitidas_generadas(self._codigo, ids, fecha_gen, eje)
        else:
            self._gestor.marcar_facturas_emitidas_generadas(self._codigo, sel, fecha_gen, self._ejercicio)
        self._view.clear_marked_ids(sel)
        self.refresh_facturas()
        self._view.show_info("Gest2A3Eco", f"Fichero generado:\n{save_path}")

    def capturar_numero_asiento_desde_a3(self):
        """
        Busca en los ficheros *A.DAT de A3ECO el número de asiento para cada
        factura seleccionada/marcada y actualiza el campo numero_asiento en BD.
        Requiere que el suenlace haya sido importado previamente por A3ECO.
        """
        if not self._ensure_write():
            return
        sel = self._view.get_marked_ids() or self._view.get_selected_ids()
        if not sel:
            self._view.show_warning("Gest2A3Eco", "Marca o selecciona al menos una factura.")
            return
        codigo_a3 = self._codigo_empresa_a3()
        actualizadas = []
        sin_asiento = []
        for fid in sel:
            fac = self._get_factura_by_id(fid)
            if not fac:
                continue
            # A3ECO almacena el num_fra de forma distinta segun el tipo de serie:
            #   - Series numericas ("1", "2"): guarda "1/0001365" (con barra)
            #   - Series alfanumericas ("A", "B"): guarda "A000031" (sin barra)
            # Construimos el formato correcto segun si la serie es un digito.
            serie = str(fac.get("serie") or "").strip()
            num_contable = self._numero_factura_contable(fac)
            if serie and serie.isdigit() and num_contable.startswith(serie):
                num_factura = f"{serie}/{num_contable[len(serie):]}"
            else:
                num_factura = num_contable
            descripcion = str(fac.get("descripcion") or "").strip()
            asiento = leer_numero_asiento_desde_a3(codigo_a3, self._ejercicio, num_factura, descripcion)
            if asiento:
                fac["numero_asiento"] = asiento
                self._gestor.upsert_factura_emitida(fac)
                actualizadas.append(f"{num_factura} → asiento {asiento}")
            else:
                sin_asiento.append(num_factura or str(fid))
        self.refresh_facturas()
        partes = []
        if actualizadas:
            partes.append("Asientos capturados:\n" + "\n".join(f"  {r}" for r in actualizadas))
        if sin_asiento:
            partes.append(
                "No encontradas en A3ECO (importa el suenlace primero):\n"
                + "\n".join(f"  {n}" for n in sin_asiento)
            )
        self._view.show_info("Gest2A3Eco", "\n\n".join(partes) if partes else "Sin cambios.")

    def enviar_a_contabilidad(self):
        """Marca las facturas seleccionadas/marcadas como pendientes en el módulo de contabilidad."""
        if not self._ensure_write():
            return
        sel = self._view.get_marked_ids() or self._view.get_selected_ids()
        if not sel:
            self._view.show_warning("Gest2A3Eco", "Marca o selecciona al menos una factura.")
            return
        self._gestor.enviar_facturas_emitidas_a_contabilidad(self._codigo, self._ejercicio, sel)
        self.refresh_facturas()
        self._view.show_info("Gest2A3Eco", f"{len(sel)} factura(s) enviadas al módulo de contabilidad.")

    def ver_asiento_factura(self):
        """Muestra el asiento contable generado por la factura seleccionada (mejora 6)."""
        from views.ui_asiento_emitida_dialog import AsientoEmitidaDialog
        sel = self._view.get_selected_ids()
        if not sel:
            self._view.show_info("Gest2A3Eco", "Selecciona una factura.")
            return
        fac = self._get_factura_by_id(sel[0])
        if not fac:
            return

        # Obtener plantilla de la factura
        plantilla = self._plantilla_emitidas_for_factura(fac, {}, set())
        ndig = int(self._empresa_conf.get("digitos_plan") or 8)

        parent_win = None
        try:
            parent_win = self._view.winfo_toplevel()
        except Exception:
            pass
        AsientoEmitidaDialog(
            parent_win,
            fac=fac,
            plantilla=plantilla,
            gestor=self._gestor,
            codigo_empresa=self._codigo,
            ndig=ndig,
            on_save=lambda _f: self.refresh_facturas(),
        )

    def _get_factura_by_id(self, fid):
        for f in self._facturas_cache or []:
            if str(f.get("id")) == str(fid):
                return f
        for f in self._listar_facturas_base():
            if str(f.get("id")) == str(fid):
                return f
        return None

    def _listar_facturas_base(self):
        if self._allow_all_years:
            return self._gestor.listar_facturas_emitidas_global(self._codigo, None)
        return self._gestor.listar_facturas_emitidas(self._codigo, self._ejercicio)

    def _year_from_factura(self, fac: dict):
        txt = str(fac.get("fecha_asiento") or fac.get("fecha_expedicion") or "").strip()
        year = self._year_from_fecha_txt(txt)
        if year is not None:
            return year
        try:
            return int(fac.get("ejercicio"))
        except Exception:
            return None

    def _year_from_fecha_txt(self, txt: str):
        txt = str(txt or "").strip()
        if not txt:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(txt, fmt).year
            except Exception:
                continue
        return None

    def _get_albaran_by_id(self, aid):
        return next(
            (
                a
                for a in self._gestor.listar_albaranes_emitidas(self._codigo, self._ejercicio)
                if str(a.get("id")) == str(aid)
            ),
            None,
        )

    def _albaranes_de_factura(self, fac: dict) -> list[dict]:
        factura_id = str(fac.get("id") or "").strip()
        eje = fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio
        if not factura_id:
            return []
        return [
            alb
            for alb in self._gestor.listar_albaranes_emitidas(self._codigo, eje)
            if str(alb.get("factura_id") or "").strip() == factura_id
        ]

    def _descripcion_factura_desde_albaranes(self, albaranes: list[dict]) -> str:
        albaranes = [a for a in (albaranes or []) if a]
        if not albaranes:
            return "Factura correspondiente al albaran."
        if len(albaranes) == 1:
            alb = albaranes[0]
            numero = str(alb.get("numero") or "").strip()
            fecha = self._format_fecha_descripcion(alb.get("fecha_asiento") or alb.get("fecha_expedicion"))
            return f"Factura correspondiente al albaran {numero} de fecha {fecha}."
        refs = []
        for alb in albaranes:
            numero = str(alb.get("numero") or "").strip()
            fecha = self._format_fecha_descripcion(alb.get("fecha_asiento") or alb.get("fecha_expedicion"))
            refs.append(f"{numero} de fecha {fecha}".strip())
        return f"Factura correspondiente a los albaranes: {'; '.join(refs)}."

    def _build_factura_lineas_desde_albaranes(self, albaranes: list[dict]) -> list[dict]:
        descripcion = self._descripcion_factura_desde_albaranes(albaranes)
        grouped = {}
        for alb in albaranes or []:
            for ln in alb.get("lineas", []) or []:
                if str(ln.get("tipo") or "").strip().lower() == "obs":
                    continue
                key = (
                    self._round4(self._to_float(ln.get("pct_iva"))),
                    self._round4(self._to_float(ln.get("pct_re"))),
                    self._round4(self._to_float(ln.get("pct_irpf"))),
                )
                item = grouped.setdefault(
                    key,
                    {
                        "concepto": descripcion,
                        "tipo": "",
                        "unidades": 1.0,
                        "precio": 0.0,
                        "base": 0.0,
                        "pct_iva": 0.0,
                        "cuota_iva": 0.0,
                        "pct_re": 0.0,
                        "cuota_re": 0.0,
                        "pct_irpf": 0.0,
                        "cuota_irpf": 0.0,
                    },
                )
                item["base"] += self._to_float(ln.get("base"))
                item["cuota_iva"] += self._to_float(ln.get("cuota_iva"))
                item["cuota_re"] += self._to_float(ln.get("cuota_re"))
                item["cuota_irpf"] += self._to_float(ln.get("cuota_irpf"))
                item["pct_iva"] = self._round4(self._to_float(ln.get("pct_iva")))
                item["pct_re"] = self._round4(self._to_float(ln.get("pct_re")))
                item["pct_irpf"] = self._round4(self._to_float(ln.get("pct_irpf")))
        out = []
        for item in grouped.values():
            item["base"] = self._round2(item["base"])
            item["cuota_iva"] = self._round2(item["cuota_iva"])
            item["cuota_re"] = self._round2(item["cuota_re"])
            item["cuota_irpf"] = self._round2(item["cuota_irpf"])
            item["precio"] = item["base"]
            out.append(item)
        return out

    def _format_fecha_descripcion(self, fecha_txt: str) -> str:
        value = str(fecha_txt or "").strip()
        if not value:
            return ""
        for fmt_in in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(value, fmt_in)
                return dt.strftime("%d/%m/%Y")
            except Exception:
                continue
        return value

    def _albaran_app_pdf_path(self, alb: dict) -> str:
        pdf_dir = str(get_default_output_dir())
        emp_name = self._safe_filename(self._empresa_conf.get("nombre") or "") or "Sin_empresa"
        safe_num = self._safe_filename(str(alb.get("numero", "") or ""))
        safe_codigo = self._safe_filename(self._codigo or "")
        safe_cliente = self._safe_filename(alb.get("nombre", "")) or "Sin_cliente"
        parts = [p for p in [f"Albaran_{safe_num}" if safe_num else "", safe_codigo, safe_cliente] if p]
        filename = "_".join(parts) if parts else f"Albaran_{alb.get('id', '')}"
        try:
            os.makedirs(os.path.join(pdf_dir, emp_name), exist_ok=True)
        except Exception:
            return ""
        return os.path.join(pdf_dir, emp_name, f"{filename}.pdf")

    def _resolve_albaran_pdf(self, alb: dict) -> str:
        pdf_path = str(alb.get("pdf_path") or "").strip()
        if pdf_path and os.path.exists(pdf_path):
            return pdf_path
        app_path = self._albaran_app_pdf_path(alb)
        if not app_path:
            return ""
        if os.path.exists(app_path):
            if pdf_path != app_path and self._can_write():
                upd = dict(alb)
                upd["pdf_path"] = app_path
                self._gestor.upsert_albaran_emitida(upd)
            return app_path
        doc = dict(alb)
        doc.setdefault("fecha_expedicion", alb.get("fecha_asiento", ""))
        doc.setdefault("fecha_operacion", alb.get("fecha_asiento", ""))
        try:
            self._generar_pdf_word(doc, app_path, default_template="albaran_template.docx")
        except Exception as e:
            self._log_pdf_error("Error al generar PDF de albaran.", e, "", app_path)
            return ""
        if self._can_write():
            upd = dict(alb)
            upd["pdf_path"] = app_path
            self._gestor.upsert_albaran_emitida(upd)
        return app_path if os.path.exists(app_path) else ""

    def _open_attachments_in_explorer(self, attachment_paths: list[str]) -> None:
        valid_paths = []
        for path in attachment_paths or []:
            norm = os.path.normpath(str(path or "").strip())
            if norm and os.path.exists(norm) and norm not in valid_paths:
                valid_paths.append(norm)
        if not valid_paths:
            return
        if len(valid_paths) == 1:
            try:
                subprocess.run(f'explorer.exe /select,"{valid_paths[0]}"', shell=True, check=False)
                return
            except Exception:
                pass
        folder = os.path.dirname(valid_paths[0])
        if folder:
            os.startfile(folder)

    def _serie(self):
        return str(self._empresa_conf.get("serie_emitidas", "A") or "A")

    def _siguiente_num(self):
        try:
            return int(self._empresa_conf.get("siguiente_num_emitidas", 1))
        except Exception:
            return 1

    def _proximo_numero(self):
        return f"{self._siguiente_num():06d}"

    def _incrementar_numeracion(self):
        self._empresa_conf["siguiente_num_emitidas"] = self._siguiente_num() + 1
        self._gestor.upsert_empresa(self._empresa_conf)

    def _listar_series_for_year(self, year: int | None, rectificativa: bool = False) -> list:
        """Devuelve lista de dicts {nombre, siguiente_num} para el desplegable de series."""
        try:
            eje = year if year is not None else self._ejercicio
            self._gestor.ensure_series_emitidas(self._codigo, eje)
            series = self._gestor.listar_series_emitidas(self._codigo, eje, es_rectificativa=1 if rectificativa else 0)
            return [s for s in series if s.get("activa", 1)]
        except Exception:
            return []

    def _serie_for_year(self, year: int | None, rectificativa: bool = False, nombre_serie: str | None = None) -> str:
        series = self._listar_series_for_year(year, rectificativa=rectificativa)
        if nombre_serie:
            match = next((s for s in series if s["nombre"] == nombre_serie), None)
            if match:
                return match["nombre"]
        if series:
            return series[0]["nombre"]
        # Fallback a empresas
        emp = self._empresa_for_year(year)
        if rectificativa:
            serie = str(emp.get("serie_emitidas_rect") or "").strip()
            return serie or "R"
        return str(emp.get("serie_emitidas", "A") or "A")

    def _siguiente_num_for_year(self, year: int | None, rectificativa: bool = False, nombre_serie: str | None = None) -> int:
        eje = year if year is not None else self._ejercicio
        serie = nombre_serie or self._serie_for_year(year, rectificativa=rectificativa)
        try:
            return self._gestor.get_siguiente_serie_num(self._codigo, eje, serie)
        except Exception:
            pass
        # Fallback
        emp = self._empresa_for_year(year)
        key = "siguiente_num_emitidas_rect" if rectificativa else "siguiente_num_emitidas"
        try:
            return int(emp.get(key, 1))
        except Exception:
            return 1

    def _proximo_numero_por_fecha(self, fecha_txt: str, rectificativa: bool = False, nombre_serie: str | None = None):
        year = self._year_from_fecha_txt(fecha_txt)
        serie = self._serie_for_year(year, rectificativa=rectificativa, nombre_serie=nombre_serie)
        num = self._siguiente_num_for_year(year, rectificativa=rectificativa, nombre_serie=serie)
        return f"{num:06d}", serie, year

    def _incrementar_numeracion_por_factura(self, fac: dict, rectificativa: bool = False):
        year = self._year_from_factura(fac)
        eje = year if year is not None else self._ejercicio
        serie_usada = str(fac.get("serie") or "").strip()
        if not serie_usada:
            serie_usada = self._serie_for_year(year, rectificativa=rectificativa)
        try:
            self._gestor.incrementar_serie_num(self._codigo, eje, serie_usada)
        except Exception:
            pass
        # Mantener compatibilidad actualizando también empresas
        emp = self._empresa_for_year(year)
        key = "siguiente_num_emitidas_rect" if rectificativa else "siguiente_num_emitidas"
        try:
            emp[key] = int(emp.get(key, 1)) + 1
        except Exception:
            emp[key] = 2
        self._gestor.upsert_empresa(emp)
        if year == self._ejercicio:
            self._empresa_conf.update(emp)

    def _ajustar_numero_por_fecha_si_aplica(self, fac: dict, numero_sugerido: str, serie_sugerida: str, rectificativa: bool = False):
        num_actual = str(fac.get("numero") or "")
        serie_actual = str(fac.get("serie") or "")
        if num_actual == str(numero_sugerido) and serie_actual == str(serie_sugerida):
            fecha_base = fac.get("fecha_asiento") or fac.get("fecha_expedicion") or ""
            sugerido, serie_sug, year = self._proximo_numero_por_fecha(fecha_base, rectificativa=rectificativa)
            fac["numero"] = sugerido
            fac["serie"] = serie_sug
            if year is not None:
                fac["ejercicio"] = year
        return fac

    def _empresa_for_year(self, year: int | None):
        if year is None:
            return dict(self._empresa_conf)
        emp = self._gestor.get_empresa(self._codigo, year)
        if emp:
            changed = False
            if not str(emp.get("serie_emitidas_rect") or "").strip():
                emp["serie_emitidas_rect"] = "R"
                changed = True
            if emp.get("siguiente_num_emitidas_rect") in (None, ""):
                emp["siguiente_num_emitidas_rect"] = 1
                changed = True
            if changed:
                self._gestor.upsert_empresa(emp)
            return emp
        base = dict(self._empresa_conf)
        base["ejercicio"] = year
        base.setdefault("serie_emitidas", "A")
        base.setdefault("siguiente_num_emitidas", 1)
        base.setdefault("serie_emitidas_rect", "R")
        base.setdefault("siguiente_num_emitidas_rect", 1)
        base["activo"] = True
        self._gestor.upsert_empresa(base)
        return base

    def _proximo_albaran_numero(self):
        pref = f"Alb-{self._ejercicio}-"
        max_n = 0
        for a in self._gestor.listar_albaranes_emitidas(self._codigo, self._ejercicio):
            num_txt = str(a.get("numero") or "")
            if pref in num_txt:
                tail = num_txt.split(pref, 1)[-1]
            else:
                tail = "".join(ch for ch in num_txt if ch.isdigit())
            digits = "".join(ch for ch in tail if ch.isdigit())
            if digits:
                try:
                    max_n = max(max_n, int(digits))
                except Exception:
                    pass
        return f"{pref}{max_n + 1:05d}"

    def _compute_total(self, fac: dict) -> float:
        total = 0.0
        lineas = aplicar_descuento_total_lineas(
            fac.get("lineas", []),
            fac.get("descuento_total_tipo"),
            fac.get("descuento_total_valor"),
        )
        for ln in lineas:
            total += (
                self._to_float(ln.get("base"))
                + self._to_float(ln.get("cuota_iva"))
                + self._to_float(ln.get("cuota_re"))
            )
        total += self._retencion_importe(fac)
        return self._round2(total)

    def _totales_factura(self, fac: dict):
        base = iva = re = 0.0
        lineas = aplicar_descuento_total_lineas(
            fac.get("lineas", []),
            fac.get("descuento_total_tipo"),
            fac.get("descuento_total_valor"),
        )
        for ln in lineas:
            base += self._to_float(ln.get("base"))
            iva += self._to_float(ln.get("cuota_iva"))
            re += self._to_float(ln.get("cuota_re"))
        ret = self._retencion_importe(fac)
        total = base + iva + re + ret
        return {
            "base": self._round2(base),
            "iva": self._round2(iva),
            "re": self._round2(re),
            "ret": self._round2(ret),
            "total": self._round2(total),
        }

    def _cliente_factura(self, fac: dict):
        fac_tercero_id = str(fac.get("tercero_id") or "").strip()
        cli = None
        if fac_tercero_id:
            cli = next(
                (t for t in self._gestor.listar_terceros() if str(t.get("id")) == fac_tercero_id),
                None,
            )
        if cli is None:
            fac_nif = normalizar_nif_cif(str(fac.get("nif") or ""))
            if fac_nif:
                cli = self._gestor.get_tercero_by_nif(fac_nif)
        if cli is None:
            cli = {}
        cp = cli.get("cp", "")
        poblacion = cli.get("poblacion", "")
        provincia = cli.get("provincia", "")
        direccion = cli.get("direccion", "")
        cp_pob = ", ".join(p for p in [cp, poblacion] if p)
        direccion_completa = "\n".join(p for p in [direccion, cp_pob, provincia] if p)
        nombre = fac.get("nombre") or cli.get("nombre", "")
        nombre_legal = cli.get("nombre_legal") or cli.get("nombre_comercial") or nombre
        pais = cli.get("pais", "") or inferir_pais_desde_identificacion(cli.get("nif") or fac.get("nif")) or "ES"
        return {
            "nombre": nombre,
            "nombre_legal": nombre_legal,
            "nombre_comercial": cli.get("nombre_comercial") or nombre,
            "nif": normalizar_nif_cif(fac.get("nif") or cli.get("nif", "")),
            "direccion": direccion,
            "cp": cp,
            "poblacion": poblacion,
            "provincia": provincia,
            "pais": pais,
            "telefono": cli.get("telefono", ""),
            "email": cli.get("email", ""),
            "contacto": cli.get("contacto", ""),
            "direccion_completa": direccion_completa,
        }

    def _cliente_relacion_factura(self, fac: dict) -> dict:
        eje = fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio
        tercero_id = str(fac.get("tercero_id") or "").strip()
        if tercero_id:
            return self._gestor.get_tercero_empresa(self._codigo, tercero_id, eje) or {}
        nif = normalizar_nif_cif(fac.get("nif"))
        if nif:
            for tercero in self._gestor.listar_terceros_por_empresa(self._codigo, eje):
                if normalizar_nif_cif(tercero.get("nif")) == nif:
                    return tercero
        return {}

    def _empresa_facturae(self) -> dict:
        emp = self._empresa_conf_for_word()
        pais = str(emp.get("pais") or "").strip().upper() or "ES"
        return {
            "nombre": emp.get("nombre", ""),
            "nombre_legal": emp.get("nombre", ""),
            "nif": normalizar_nif_cif(emp.get("cif", "")),
            "direccion": emp.get("direccion", ""),
            "cp": emp.get("cp", ""),
            "poblacion": emp.get("poblacion", ""),
            "provincia": emp.get("provincia", ""),
            "pais": pais,
            "telefono": emp.get("telefono", ""),
            "email": emp.get("email", ""),
        }

    def _factura_to_rows(self, fac: dict):
        rows = []
        base_row = {
            "Serie": fac.get("serie", ""),
            "Numero Factura": self._numero_factura_contable(fac),
            "Numero Factura Largo SII": fac.get("numero_largo_sii", ""),
            "Fecha Asiento": fac.get("fecha_asiento", ""),
            "Fecha Expedicion": fac.get("fecha_expedicion") or fac.get("fecha_asiento", ""),
            "Fecha Operacion": fac.get("fecha_operacion", ""),
            "Tipo Operacion": fac.get("tipo_operacion", "01"),
            "Impreso": fac.get("modelo_fiscal", ""),
            "Descripcion Factura": fac.get("descripcion", ""),
            "NIF Cliente Proveedor": normalizar_nif_cif(fac.get("nif", "")),
            "Nombre Cliente Proveedor": fac.get("nombre", ""),
            "_pdf_ref": self._pdf_ref_base(fac.get("pdf_ref", "")),
        }
        subcuenta_cliente = self._resolved_subcuenta_cliente(fac)
        if subcuenta_cliente:
            base_row["Cuenta Cliente Proveedor"] = subcuenta_cliente
        lineas = aplicar_descuento_total_lineas(
            list(fac.get("lineas", [])),
            fac.get("descuento_total_tipo"),
            fac.get("descuento_total_valor"),
        )
        ret_pct = self._round2(self._to_float(fac.get("retencion_pct")))
        ret_importe = self._retencion_importe(fac)
        lineas_validas = [ln for ln in lineas if str(ln.get("tipo") or "").strip().lower() != "obs"]
        for idx, ln in enumerate(lineas_validas):
            aplica_ret_pct = ret_pct != 0
            ret_aplica_linea = (ret_importe != 0 and idx == len(lineas_validas) - 1) if not aplica_ret_pct else True
            r = dict(base_row)
            r.update(
                {
                    "Descripcion Linea": ln.get("concepto", "") or fac.get("descripcion", ""),
                    "Base": self._round2(self._to_float(ln.get("base"))),
                    "Cuota IVA": self._round2(self._to_float(ln.get("cuota_iva"))),
                    "Porcentaje IVA": self._round2(self._to_float(ln.get("pct_iva"))),
                    "Porcentaje Recargo Equivalencia": self._round2(self._to_float(ln.get("pct_re"))),
                    "Cuota Recargo Equivalencia": self._round2(self._to_float(ln.get("cuota_re"))),
                    "Porcentaje Retencion IRPF": ret_pct if aplica_ret_pct else 0.0,
                    "Cuota Retencion IRPF": 0.0 if aplica_ret_pct else (ret_importe if ret_aplica_linea else 0.0),
                }
            )
            rows.append(r)
        return rows

    def _numero_factura_contable(self, fac: dict) -> str:
        serie = str(fac.get("serie") or "").strip()
        numero = str(fac.get("numero") or "").strip()
        if not serie:
            return numero
        if not numero:
            return serie
        # Evitar duplicar la serie si 'numero' ya la lleva como prefijo (ej. "A000001" + serie "A")
        if numero.upper().startswith(serie.upper()):
            numero = numero[len(serie):]
        return f"{serie}{numero}"

    def _docx_template_path(
        self,
        fac: dict | None = None,
        warn_missing: bool = False,
        default_filename: str = "factura_emitida_template.docx",
    ) -> str:
        tpl_dir = get_word_templates_dir()
        default_path = os.path.join(tpl_dir, default_filename)
        if not fac:
            return default_path
        chosen = str(fac.get("plantilla_word") or "").strip()
        if not chosen:
            return default_path
        candidate = os.path.join(tpl_dir, chosen)
        if os.path.exists(candidate):
            return candidate
        if warn_missing:
            self._view.show_warning(
                "Gest2A3Eco",
                f"No se encuentra la plantilla seleccionada:\n{chosen}\nSe usara la plantilla por defecto.",
            )
        return default_path

    def _default_plantilla_emitidas_name(self, ejercicio: int | None) -> str:
        eje = ejercicio if ejercicio is not None else self._ejercicio
        plantillas = self._gestor.listar_emitidas(self._codigo, eje)
        if not plantillas:
            return ""
        return str(plantillas[0].get("nombre") or "")

    def _resolved_subcuenta_cliente(self, fac: dict) -> str:
        eje = fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio
        tercero_id = str(fac.get("tercero_id") or "").strip()
        if tercero_id:
            rel = self._gestor.get_tercero_empresa(self._codigo, tercero_id, eje) or {}
            subcuenta = str(rel.get("subcuenta_cliente") or "").strip()
            if subcuenta:
                return subcuenta
        nif = normalizar_nif_cif(fac.get("nif"))
        if nif:
            for tercero in self._gestor.listar_terceros_por_empresa(self._codigo, eje):
                if normalizar_nif_cif(tercero.get("nif")) == nif:
                    subcuenta = str(tercero.get("subcuenta_cliente") or "").strip()
                    if subcuenta:
                        return subcuenta
                    break
        return str(fac.get("subcuenta_cliente") or "").strip()

    def _plantilla_emitidas_tipo(self) -> dict:
        return {
            "nombre": "__emitidas_tipo__",
            "codigo_empresa": self._codigo,
            "ejercicio": self._ejercicio,
            "cuenta_cliente_prefijo": "430",
            "cuenta_ingreso_por_defecto": "70000000",
            "cuenta_iva_repercutido_defecto": "47700000",
            "cuenta_retenciones_irpf": "",
            "excel": {},
        }

    def _plantilla_emitidas_for_factura(self, fac: dict, cache: dict, no_tpl_years: set) -> dict:
        eje = fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio
        if eje not in cache:
            cache[eje] = self._gestor.listar_emitidas(self._codigo, eje)
        plantillas = cache[eje]
        nombre_plantilla = str(fac.get("plantilla_emitidas") or "").strip()
        if plantillas:
            if nombre_plantilla:
                for plantilla in plantillas:
                    if str(plantilla.get("nombre") or "").strip() == nombre_plantilla:
                        return plantilla
            return plantillas[0]
        no_tpl_years.add(eje)
        plantilla_tipo = self._plantilla_emitidas_tipo()
        plantilla_tipo["ejercicio"] = eje
        return plantilla_tipo

    def _empresa_conf_for_word(self) -> dict:
        # Lee datos frescos de la BD para garantizar que el PDF usa la config actual
        fresh = self._gestor.get_empresa(self._codigo, self._ejercicio) or {}
        # Si la config fresca no tiene domicilio, intentar con el ejercicio mas reciente
        if not any(fresh.get(f) for f in ("direccion", "cp", "poblacion", "provincia")):
            fallback = self._gestor.get_empresa(self._codigo, None) or {}
            for f in ("direccion", "cp", "poblacion", "provincia", "cif", "telefono", "email", "logo_path",
                      "logo_max_width_mm", "logo_max_height_mm"):
                if not fresh.get(f) and fallback.get(f):
                    fresh[f] = fallback[f]
        emp = dict(self._empresa_conf or {})
        emp.update(fresh)
        emp.setdefault("codigo", self._codigo)
        emp.setdefault("codigo_empresa", self._codigo)
        return emp

    def _persist_factura_if_allowed(self, factura: dict) -> None:
        if not self._can_write():
            return
        self._gestor.upsert_factura_emitida(factura)

    def _store_pdf_copies(self, fac: dict, src_path: str) -> None:
        fac = self._ensure_pdf_ref(fac)
        upd = dict(fac)
        app_path = self._app_pdf_path(fac)
        if app_path:
            try:
                shutil.copy2(src_path, app_path)
                upd["pdf_path"] = app_path
            except Exception:
                pass
        self._persist_factura_if_allowed(upd)

    def _generar_pdf_word(self, fac: dict, out_path: str, default_template: str | None = None) -> None:
        """Genera un PDF usando siempre la plantilla Word. Lanza excepcion si no hay plantilla o falla."""
        template_path = self._docx_template_path(fac, warn_missing=False, default_filename=default_template or "factura_emitida_template.docx")
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"No se encuentra la plantilla Word:\n{template_path}")
        cliente = self._cliente_factura(fac)
        tot = self._totales_factura(fac)
        context = build_context_emitida(self._empresa_conf_for_word(), fac, cliente, tot)
        generar_pdf_desde_plantilla_word(
            template_path=template_path,
            context=context,
            out_pdf_path=out_path,
            guardar_docx=False,
        )

    def _resolve_app_pdf(self, fac: dict) -> str:
        """Devuelve la ruta del PDF en la carpeta de la app, generandolo si no existe o esta obsoleto.

        Mejora 3: si la factura tiene updated_at mas reciente que el fichero PDF, el PDF
        se considera obsoleto y se regenera automaticamente.
        """
        import logging as _logging
        _pdf_log = _logging.getLogger(__name__)

        app_path = self._app_pdf_path(fac)
        if not app_path:
            return ""

        if os.path.exists(app_path):
            # Comprobar si la factura fue modificada despues de que se genero el PDF
            fac_updated = str(fac.get("updated_at") or "").strip()
            pdf_obsoleto = False
            if fac_updated:
                try:
                    pdf_mtime = datetime.fromtimestamp(os.path.getmtime(app_path))
                    # updated_at puede incluir 'T', 'Z' u offset; normalizar a YYYY-MM-DD HH:MM:SS
                    fac_str = fac_updated.replace("T", " ").replace("Z", "").strip()[:19]
                    fac_dt = datetime.strptime(fac_str, "%Y-%m-%d %H:%M:%S")
                    if fac_dt > pdf_mtime:
                        pdf_obsoleto = True
                        _pdf_log.info(
                            "PDF obsoleto para factura id=%s (updated_at=%s > pdf_mtime=%s);"
                            " regenerando: %s",
                            fac.get("id"), fac_dt, pdf_mtime, app_path,
                        )
                except Exception:
                    pass  # Si no podemos comparar, usamos el PDF existente

            # Comprobar si la config de empresa cambio despues de que se genero el PDF
            if not pdf_obsoleto:
                try:
                    emp_fresh = self._gestor.get_empresa(self._codigo, self._ejercicio) or {}
                    emp_cached = self._empresa_conf or {}
                    _addr_fields = ("direccion", "cp", "poblacion", "provincia", "nombre", "cif")
                    if any(str(emp_fresh.get(f) or "") != str(emp_cached.get(f) or "") for f in _addr_fields):
                        pdf_obsoleto = True
                        _pdf_log.info(
                            "PDF obsoleto para factura id=%s: config empresa ha cambiado; regenerando: %s",
                            fac.get("id"), app_path,
                        )
                except Exception:
                    pass

            if pdf_obsoleto:
                try:
                    os.unlink(app_path)
                except Exception as _e:
                    _pdf_log.warning("No se pudo eliminar PDF obsoleto %s: %s", app_path, _e)
            else:
                return app_path

        try:
            self._generar_pdf_word(fac, app_path)
        except Exception as e:
            self._log_pdf_error("Error al generar PDF.", e, "", app_path)
            return ""
        if os.path.exists(app_path):
            upd = dict(fac)
            upd["pdf_path"] = app_path
            self._persist_factura_if_allowed(upd)
            return app_path
        return ""

    def _ensure_app_pdf(self, fac: dict) -> None:
        pdf_ref = self._pdf_ref_base(fac.get("pdf_ref") or "")
        if not pdf_ref:
            return
        app_path = self._app_pdf_path(fac)
        if not app_path:
            return
        if os.path.exists(app_path):
            if fac.get("pdf_path") != app_path:
                upd = dict(fac)
                upd["pdf_path"] = app_path
                self._persist_factura_if_allowed(upd)
            return
        existing_path = str(fac.get("pdf_path") or "").strip()
        if existing_path and os.path.exists(existing_path):
            try:
                shutil.copy2(existing_path, app_path)
                upd = dict(fac)
                if existing_path.lower().startswith("z:\\") and not fac.get("pdf_path_a3"):
                    upd["pdf_path_a3"] = existing_path
                upd["pdf_path"] = app_path
                self._persist_factura_if_allowed(upd)
                return
            except Exception:
                pass
        a3_path = str(fac.get("pdf_path_a3") or "").strip()
        if a3_path and os.path.exists(a3_path):
            try:
                shutil.copy2(a3_path, app_path)
                upd = dict(fac)
                upd["pdf_path"] = app_path
                self._persist_factura_if_allowed(upd)
                return
            except Exception:
                pass
        try:
            self._generar_pdf_word(fac, app_path)
        except Exception as e:
            self._log_pdf_error("_ensure_app_pdf: error Word.", e, "", app_path)
            return
        upd = dict(fac)
        upd["pdf_path"] = app_path
        self._persist_factura_if_allowed(upd)

    def _ensure_a3_pdf(self, fac: dict) -> None:
        import logging as _logging
        _a3_log = _logging.getLogger(__name__)
        pdf_ref = self._pdf_ref_base(fac.get("pdf_ref") or "")
        if not pdf_ref:
            _a3_log.warning("_ensure_a3_pdf: factura id=%s sin pdf_ref, saltada.", fac.get("id"))
            return
        a3_path = self._a3_pdf_path_for(pdf_ref, fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio)
        if not a3_path:
            _a3_log.warning("_ensure_a3_pdf: factura id=%s pdf_ref=%s -> a3_path vacio (Z: no accesible?).", fac.get("id"), pdf_ref)
            return
        if os.path.exists(a3_path):
            if fac.get("pdf_path_a3") != a3_path:
                upd = dict(fac)
                upd["pdf_path_a3"] = a3_path
                self._persist_factura_if_allowed(upd)
            return
        app_path = str(fac.get("pdf_path") or "").strip()
        if app_path and os.path.exists(app_path):
            try:
                shutil.copy2(app_path, a3_path)
                upd = dict(fac)
                upd["pdf_path_a3"] = a3_path
                self._persist_factura_if_allowed(upd)
                return
            except Exception as e:
                _a3_log.warning("_ensure_a3_pdf: error copiando desde app_path para id=%s: %s", fac.get("id"), e)
        try:
            self._generar_pdf_word(fac, a3_path)
        except Exception as e:
            self._log_pdf_error("_ensure_a3_pdf: error Word.", e, "", a3_path)
            return
        upd = dict(fac)
        upd["pdf_path_a3"] = a3_path
        self._persist_factura_if_allowed(upd)

    def _app_pdf_path(self, fac: dict) -> str:
        pdf_dir = str(get_default_output_dir())
        emp_name = self._safe_filename(self._empresa_conf.get("nombre") or "") or "Sin_empresa"
        serie = self._safe_filename(str(fac.get("serie", "") or ""))
        num = self._safe_filename(str(fac.get("numero", "") or ""))
        codigo = self._safe_filename(self._codigo or "")
        cliente = self._safe_filename(fac.get("nombre", "")) or "Sin_cliente"
        id_part = f"{serie}{num}" if (serie or num) else ""
        parts = [p for p in [id_part, codigo, cliente] if p]
        filename = "_".join(parts) if parts else f"Factura_{fac.get('id', '')}"
        try:
            os.makedirs(os.path.join(pdf_dir, emp_name), exist_ok=True)
        except Exception:
            return ""
        if not filename:
            return ""
        return os.path.join(pdf_dir, emp_name, f"{filename}.pdf")

    def _a3_pdf_path(self, pdf_ref: str) -> str:
        return self._a3_pdf_path_for(pdf_ref, self._ejercicio)

    def _a3_pdf_path_for(self, pdf_ref: str, ejercicio) -> str:
        codigo = self._codigo_empresa_a3()
        ejercicio = str(ejercicio or "").strip()
        if not codigo or not ejercicio:
            return ""
        base_dir = os.path.join("Z:\\", "A3", "A3ECO", codigo, "FACTURAS", ejercicio)
        try:
            os.makedirs(base_dir, exist_ok=True)
        except Exception:
            return ""
        ref = self._pdf_ref_base(pdf_ref)
        filename = f"{ref}.pdf" if ref else ""
        if not filename:
            return ""
        return os.path.join(base_dir, filename)

    def _ensure_pdf_ref(self, fac: dict) -> dict:
        ref = self._pdf_ref_base(fac.get("pdf_ref") or "")
        if ref:
            if str(fac.get("pdf_ref") or "").strip() != ref:
                fac = dict(fac)
                fac["pdf_ref"] = ref
                self._persist_factura_if_allowed(fac)
            return fac
        ref = str(
            self._gestor.next_pdf_ref(
                self._codigo,
                fac.get("ejercicio") if fac.get("ejercicio") is not None else self._ejercicio,
            )
        ).strip()
        fac = dict(fac)
        fac["pdf_ref"] = ref
        self._persist_factura_if_allowed(fac)
        return fac

    def _prepare_facturas_for_suenlace(self, facturas: list[dict]) -> list[dict]:
        prepared = []
        selected_counts = {}
        for fac in facturas:
            base = self._pdf_ref_base(fac.get("pdf_ref") or "")
            if base:
                selected_counts[base] = selected_counts.get(base, 0) + 1

        all_facturas = self._listar_facturas_base()
        global_counts = {}
        for fac in all_facturas:
            base = self._pdf_ref_base(fac.get("pdf_ref") or "")
            if base:
                global_counts[base] = global_counts.get(base, 0) + 1

        assigned = set()
        for fac in facturas:
            current = self._pdf_ref_base(fac.get("pdf_ref") or "")
            needs_new_ref = (
                not current
                or selected_counts.get(current, 0) > 1
                or global_counts.get(current, 0) > 1
                or current in assigned
            )
            if needs_new_ref:
                fac = dict(fac)
                fac["pdf_ref"] = ""
            fac = self._ensure_pdf_ref(fac)
            assigned.add(self._pdf_ref_base(fac.get("pdf_ref") or ""))
            self._ensure_app_pdf(fac)
            # Despues de ensure_app_pdf el dict fac no se actualiza en memoria,
            # pero el PDF ya existe en disco. Actualizamos pdf_path en el dict
            # para que ensure_a3_pdf pueda copiar desde ahi en vez de regenerar.
            # Se actualiza siempre que el fichero exista en app_path y el dict
            # apunte a una ruta distinta (o vacia/obsoleta), evitando que una
            # ruta antigua heredada bloquee la copia al directorio de A3ECO.
            app_path = self._app_pdf_path(fac)
            if app_path and os.path.exists(app_path) and fac.get("pdf_path") != app_path:
                fac = dict(fac)
                fac["pdf_path"] = app_path
            self._ensure_a3_pdf(fac)
            prepared.append(fac)
        return prepared

    def _pdf_ref_base(self, ref: str) -> str:
        value = str(ref or "").strip()
        if not value:
            return ""
        return value.split("@", 1)[0].strip()

    def _log_pdf_error(self, msg: str, exc: Exception, template_path: str, save_path: str) -> None:
        try:
            log_path = get_log_path("pdf_error.log")
            with open(log_path, "a", encoding="utf-8") as f:
                f.write("\n---- PDF ERROR ----\n")
                f.write(f"Message: {msg}\n")
                f.write(f"Template: {template_path}\n")
                f.write(f"Output: {save_path}\n")
                f.write("Exception:\n")
                f.write("".join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
        except Exception:
            pass

    def _to_float(self, x) -> float:
        try:
            if x is None or x == "":
                return 0.0
            if isinstance(x, (int, float)) and not isinstance(x, bool):
                return float(x)
            s = str(x).strip().replace("\\xa0", " ")
            if "." in s and "," in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            return float(s)
        except Exception:
            return 0.0

    def _retencion_importe(self, fac: dict) -> float:
        if not fac:
            return 0.0
        if not bool(fac.get("retencion_aplica")):
            ret_lineas = 0.0
            for ln in fac.get("lineas", []):
                ret_lineas += self._to_float(ln.get("cuota_irpf"))
            return self._round2(ret_lineas)
        signo_ret = self._signo_retencion_por_factura(fac)
        importe = fac.get("retencion_importe")
        if importe is None or importe == "":
            base_raw = fac.get("retencion_base")
            if base_raw is None or base_raw == "":
                base = self._base_imponible_descuento(fac)
            else:
                base = self._to_float(base_raw)
            pct = self._to_float(fac.get("retencion_pct"))
            return self._round2(signo_ret * abs(base * pct / 100.0)) if pct else 0.0
        return self._round2(signo_ret * abs(self._to_float(importe)))

    def _signo_retencion_por_factura(self, fac: dict) -> float:
        base_raw = fac.get("retencion_base")
        if base_raw is None or base_raw == "":
            base = self._base_imponible_descuento(fac)
        else:
            base = self._to_float(base_raw)
        return 1.0 if base < 0 else -1.0

    def _base_imponible_descuento(self, fac: dict) -> float:
        total = 0.0
        lineas = aplicar_descuento_total_lineas(
            fac.get("lineas", []),
            fac.get("descuento_total_tipo"),
            fac.get("descuento_total_valor"),
        )
        for ln in lineas:
            if str(ln.get("tipo") or "").strip().lower() == "obs":
                continue
            total += self._to_float(ln.get("base"))
        return self._round2(total)

    def _negate_factura_lineas(self, lineas):
        out = []
        for ln in lineas or []:
            nl = dict(ln)
            if str(nl.get("tipo") or "").strip().lower() == "obs":
                out.append(nl)
                continue
            nl["base"] = self._negate_value(nl.get("base"))
            nl["cuota_iva"] = self._negate_value(nl.get("cuota_iva"))
            nl["cuota_re"] = self._negate_value(nl.get("cuota_re"))
            nl["cuota_irpf"] = self._negate_value(nl.get("cuota_irpf"))
            out.append(nl)
        return out

    def _negate_value(self, value):
        if value in (None, ""):
            return value
        try:
            val = self._to_float(value)
        except Exception:
            return value
        if val == 0:
            return 0.0
        return -abs(val)

    def _safe_filename(self, value: str) -> str:
        s = (value or "").strip()
        if not s:
            return ""
        bad = '<>:"/\\\\|?*'
        for ch in bad:
            s = s.replace(ch, " ")
        s = " ".join(s.split())
        return s

    def _cuenta_bancaria_default(self) -> str:
        raw = self._empresa_conf.get("cuenta_bancaria") or ""
        if str(raw).strip():
            return str(raw).strip()
        cuentas = self._empresa_conf.get("cuentas_bancarias") or ""
        for sep in ["\n", ";", ","]:
            cuentas = str(cuentas).replace(sep, ",")
        for p in str(cuentas).split(","):
            p = p.strip()
            if p:
                return p
        return ""

    def _round2(self, x) -> float:
        try:
            return round(float(x), 2)
        except Exception:
            return 0.0

    def _round4(self, x) -> float:
        try:
            return round(float(x), 4)
        except Exception:
            return 0.0

    # ------------------- Exportacion Excel -------------------

    def exportar_facturas_excel(self, fecha_desde: str, fecha_hasta: str, cliente_filter: str):
        from tkinter import filedialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from utils.ui_facturas_emitidas_helpers import parse_date_ui

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"facturas_{self._ejercicio}.xlsx",
            parent=self._view,
        )
        if not filepath:
            return

        desde = None
        hasta = None
        try:
            if fecha_desde:
                desde = parse_date_ui(fecha_desde)
        except Exception:
            pass
        try:
            if fecha_hasta:
                hasta = parse_date_ui(fecha_hasta)
        except Exception:
            pass

        rows = []
        for fac in self._facturas_cache:
            if cliente_filter:
                nombre = (fac.get("nombre") or "").lower()
                nif = (fac.get("nif") or "").lower()
                if cliente_filter not in nombre and cliente_filter not in nif:
                    continue
            if desde or hasta:
                try:
                    fd = parse_date_ui(fac.get("fecha_asiento", ""))
                    if desde and fd < desde:
                        continue
                    if hasta and fd > hasta:
                        continue
                except Exception:
                    pass
            rows.append(fac)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"
        headers = ["Ejercicio", "Serie", "Numero", "Fecha", "Cliente", "NIF", "Base", "IVA", "RE", "Total", "Generada"]
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="2563EB")
        hdr_align = Alignment(horizontal="center")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
        for fac in rows:
            tots = self._totales_factura(fac)
            ws.append([
                fac.get("ejercicio", ""),
                str(fac.get("serie") or "").strip(),
                str(fac.get("numero") or "").strip(),
                fac.get("fecha_asiento", ""),
                fac.get("nombre", ""),
                fac.get("nif", ""),
                tots["base"],
                tots["iva"],
                tots["re"],
                tots["total"],
                "Si" if fac.get("generada") else "No",
            ])
        col_widths = [10, 8, 14, 12, 35, 16, 12, 12, 10, 12, 10]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
        wb.save(filepath)
        self._view.show_info("Gest2A3Eco", f"Exportadas {len(rows)} facturas.")

    def exportar_albaranes_excel(self, fecha_desde: str, fecha_hasta: str, cliente_filter: str):
        from tkinter import filedialog
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from utils.ui_facturas_emitidas_helpers import parse_date_ui

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"albaranes_{self._ejercicio}.xlsx",
            parent=self._view,
        )
        if not filepath:
            return

        desde = None
        hasta = None
        try:
            if fecha_desde:
                desde = parse_date_ui(fecha_desde)
        except Exception:
            pass
        try:
            if fecha_hasta:
                hasta = parse_date_ui(fecha_hasta)
        except Exception:
            pass

        albaranes = self._gestor.listar_albaranes_emitidas(self._codigo, self._ejercicio)
        facturas_map = {
            str(f.get("id")): str(f.get("numero") or "").strip()
            for f in self._gestor.listar_facturas_emitidas(self._codigo, self._ejercicio)
        }

        rows = []
        for alb in albaranes:
            if cliente_filter:
                nombre = (alb.get("nombre") or "").lower()
                nif = (alb.get("nif") or "").lower()
                if cliente_filter not in nombre and cliente_filter not in nif:
                    continue
            if desde or hasta:
                try:
                    fd = parse_date_ui(alb.get("fecha_asiento", ""))
                    if desde and fd < desde:
                        continue
                    if hasta and fd > hasta:
                        continue
                except Exception:
                    pass
            rows.append(alb)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Albaranes"
        headers = ["Numero", "Fecha", "Cliente", "NIF", "Base", "IVA", "RE", "Total", "Facturado", "Factura"]
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="2563EB")
        hdr_align = Alignment(horizontal="center")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
        for alb in rows:
            tots = self._totales_factura(alb)
            factura_txt = facturas_map.get(str(alb.get("factura_id") or ""), "")
            ws.append([
                str(alb.get("numero") or "").strip(),
                alb.get("fecha_asiento", ""),
                alb.get("nombre", ""),
                alb.get("nif", ""),
                tots["base"],
                tots["iva"],
                tots["re"],
                tots["total"],
                "Si" if alb.get("facturado") else "No",
                factura_txt,
            ])
        col_widths = [18, 12, 35, 16, 12, 12, 10, 12, 10, 14]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, ci).column_letter].width = w
        wb.save(filepath)
        self._view.show_info("Gest2A3Eco", f"Exportados {len(rows)} albaranes.")
